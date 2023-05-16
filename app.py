from flask import Flask, render_template
from openpyxl import load_workbook
import os
from openpyxl.styles import Font

app = Flask(__name__, static_folder='/Users/anijahphillip/Desktop/ReportApp/static')

@app.route('/')
def display_workbook():
    # Load the workbook
    workbook = load_workbook('/Users/anijahphillip/Downloads/Security_Awarenessreport.xlsx')

    # Extract the sheet names
    sheet_names = workbook.sheetnames

    # Apply formatting to the pivot table in each sheet
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        max_row = sheet.max_row

        # Apply bold font to the left column (A column)
        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(bold=True)

        # Apply larger font size to the title (A1 cell)
        title_cell = sheet['A1']
        title_cell.font = Font(size=14)

    base_dir = os.path.dirname(os.path.abspath(__file__))

    # Render the template and pass the sheet names and workbook data
    return render_template('workbook.html', sheet_names=sheet_names, workbook=workbook)

if __name__ == '__main__':
    app.run(port=8080, debug=True)
