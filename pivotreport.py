import os
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import PieChart3D, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import matplotlib.pyplot as plt

# Load the raw data from Excel file into a pandas DataFrame
df = pd.read_excel('/Users/anijahphillip/Downloads/reportdata.xlsx', skiprows=6)

# Get a list of all unique agencies in the data
agencies = df['AGENCY'].unique()

# Create a new workbook
workbook = Workbook()
workbook.remove(workbook.active)


# Loop through each agency and create a worksheet with a pivot table and pie chart
for agency in agencies:
    # Filter the data to include only the current agency
    agency_data = df[df['AGENCY'] == agency]

    # Create a pivot table for the current agency
    pivot_table = pd.pivot_table(agency_data, values='AGENCY', index='02 - Information Security in the Workplace', aggfunc='count', margins=True, margins_name='Grand Total')
    pivot_table.index = pivot_table.index.rename('')
    pivot_table = pivot_table.rename(columns={'AGENCY': agency.upper()})

    worksheet = workbook.create_sheet(title=agency.upper())

    for r in dataframe_to_rows(pivot_table, index=True, header=True):
        worksheet.append(r)

    pie_chart = PieChart3D()
    labels = Reference(worksheet, min_col=1, min_row=3, max_row=5)
    data = Reference(worksheet, min_col=2, max_col=pivot_table.shape[1]+1, min_row=2, max_row=pivot_table.shape[0]+1)
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(labels)
    pie_chart.title = f'{agency.upper()} % Rate'
    worksheet.add_chart(pie_chart, f"C{pivot_table.shape[0]+3}")

    # Convert the data and labels to Python lists
    labels = pivot_table.index.tolist()
    data = pivot_table.iloc[:, 0].tolist()


    fig, ax = plt.subplots()
    ax.pie(data, labels=labels, autopct='%1.1f%%')
    ax.set_title(f'{agency.upper()} % Rate')

    # Save the chart as an image
    chart_filename = f"{agency.upper()}_pie_chart.png"
    chart_filepath = os.path.join('/Users/anijahphillip/Desktop/ReportApp/static/PieCharts', chart_filename)
    plt.savefig(chart_filepath)
    plt.close()
  
    # for sheet_name in workbook.sheetnames:
    #     sheet = workbook[agency]

    # for chart in sheet._charts:
    #     # Extract chart information or save it as an image
    #     chart_filename = f"{agency.upper()}_pie_chart.png"
    #     chart_filepath = os.path.join('/Users/anijahphillip/Desktop/PieCharts', chart_filename)
    #     chart.save(chart_filepath)
            
                    # Save the chart as PNG


        

# Save the workbook to an Excel file
output_file = '/Users/anijahphillip/Downloads/Security_Awarenessreport.xlsx'
workbook.save(output_file)
